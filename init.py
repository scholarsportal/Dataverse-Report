#The following program creates an excel file with worksheets showing monthly totals over
#1. downloads by dataverse (high level)
#2. downloads by dataset (lower level)
#3. datatypes loaded
#4. dataset subjects
#5. accounts created by affiliation

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import time
import sql_connect
######

with open('sql/get_objects.sql') as f:
	get_objects_sql = f.read()
with open('sql/get_dataverse_name.sql') as f:
	get_dataverse_name_sql = f.read()
with open('sql/downloads_by_month.sql') as f:
	downloads_by_month_sql = f.read()

##SHEET 2
with open('sql/get_datasets.sql') as f:
	get_datasets_sql = f.read()
with open('sql/get_dataset_parents.sql') as f:
	get_dataset_parents_sql = f.read()
with open('sql/dataset_downloads_by_month.sql') as f:
	dataset_downloads_by_month_sql = f.read()
with open('sql/get_dataset_name.sql') as f:
	get_dataset_name_sql = f.read()
with open('sql/get_dataset_files.sql')as f:
	get_dataset_files_sql = f.read()
with open('sql/get_dataset_status.sql')as f:
	get_dataset_status_sql = f.read()

##SHEET 3
with open('sql/get_content_types.sql') as f:
	get_content_types_sql = f.read()

##SHEET 4
with open('sql/get_subjects.sql') as f:
	get_subjects_sql = f.read()
with open('sql/get_dataset_subjects.sql') as f:
	get_dataset_subjects_sql = f.read()

##SHEET 5
with open('sql/get_affiliations.sql') as f:
	get_affiliations_sql = f.read()
with open('sql/get_users_by_month_affiliations.sql') as f:
	get_users_by_month_affiliations_sql = f.read()

#this is a recursive function allowing the drilldown of relationships from whichever level you start with - determined by dataverse_id 
def getSubDataverses(dataverse_id,level):
	#dataverse_id - the id of the root dataverse
	#level - the current level of the root dataverse
	global table_header
	cur = conn.cursor()
	cur.execute(get_objects_sql.replace("{dataverse_id}", str(dataverse_id)))
	rows = cur.fetchall()
	#
	for row in rows:
		table_row=[]#reset the table row so it can be added with new data
		for r in row:
			table_row.append(r)
		name=""
		#
		if table_row[1]=="Dataverse":
			info = getObjectName(table_row[0],get_dataverse_name_sql)
			name=info[0]
			array= normalizeArray(level,info[1],name, table_row[2])
			getDownloadsByMonthResult(table_row[0],array,downloads_by_month_sql)
		#check if we should go deeper	
		if level<level_count:	
			getSubDataverses(table_row[0],level+1)
############3
def getDatasets(route):
	cur = conn.cursor()
	cur.execute(get_datasets_sql)
	rows = cur.fetchall()
	#
	for row in rows:
		table_row=[]#reset the table row so it can be added with new data
		for r in row:
			table_row.append(r)
		info = getObjectName(table_row[0],get_dataset_name_sql)
		title=info[0]
		name=info[1]
		status =getDatasetStatus(table_row[0])
		hierarchy=list(reversed(getHierarchy(table_row[0])))
		del hierarchy[0] # remove first item from list
		
		if len(hierarchy) is 0:
			root="Root"
		else:
			root = hierarchy[0]
			del hierarchy[0]
		array=[root,">".join(hierarchy),title,name,"Dataset"]
		array.append(status)
		array.append(table_row[2])

		#
		files_array=getDatasetFiles(table_row[0],1)
		try:
			array.append(round(int(sum(files_array))/1024,1))#convert to KB
		except:
			array.append(0)#when NoneType experienced
		if route is "BY_MONTH":
			getDownloadsByMonthResult(table_row[0],array,dataset_downloads_by_month_sql)
		if route is "BY_CONTENTTYPE":
			getDatasetContent(table_row[0],array)	
		if route is "BY_SUBJECT":
			getDatasetSubject(table_row[0],array)	

def getDatasetStatus(dataset_id):
	cur = conn.cursor()
	cur.execute(get_dataset_status_sql.replace("{id}", str(dataset_id)))
	rows = cur.fetchall()
	for r in rows:
		return r[0]


def getHierarchy(object_id):
	cur = conn.cursor()
	cur.execute(get_dataset_parents_sql.replace("{object_id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for r in rows:
		array.append(getObjectName(r[0],get_dataverse_name_sql)[0])
	return array
#########			
def normalizeArray(level,type,name,date):
	_table_row = [None] * (level-1)
	_table_row.append(name)
	#Also append blank cols to the end
	padding = [None] * (level_count-level)
	array = np.concatenate((_table_row, padding), axis=0).tolist()
	array.append(type)
	array.append(date)
	return array

def getObjectName(object_id,sql):
	cur = conn.cursor()
	cur.execute(sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	return [rows[0][0],rows[0][1]]
def getDownloadsByMonth(object_id,sql):
	global start_date,end_date
	return sql.replace("{object_id}", str(object_id)).replace("{start_date}", start_date).replace("{end_date}", end_date)

def getDownloadsByMonthResult(object_id,_table_row,sql):
	global needs_table_header,ws_row_count,ws_col_start,ws_cols_count
	# global table_header
	cur = conn.cursor()
	cur.execute(getDownloadsByMonth(object_id,sql))
	rows = cur.fetchall()
	for row in rows:
		#take the first value and add it to the table_header
		#take the second value and add it to the table_row
		if needs_table_header:
			table_header.append(row[0].strftime("%b")+" - "+row[0].strftime("%Y"))
		r =	row[1]
		if r is None:
   			 r = 0	
		_table_row.append(r)
		#	
	if needs_table_header:
		needs_table_header=False
		table_header.append("Total")#add total col
		ws.append(table_header)#first line of workbook
	ws_row_count+=1
	ws_col_end=get_column_letter(ws_col_start+ws_cols_count)
	_table_row.append("=SUM("+get_column_letter(ws_col_start)+str(ws_row_count)+":"+ws_col_end+str(ws_row_count)+")")#add calculated cell
	ws.append(_table_row)#second to n line of workbook
def getDatasetTypes():
	cur = conn.cursor()
	cur.execute(get_content_types_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[0])
	return array
def getDatasetContent(object_id,_table_row):
	global contenttypes,ws_col_start,ws_row_count
	files_array=getDatasetFiles(object_id,0)
	for i in range(len(contenttypes)):
		_table_row.append(files_array.count(contenttypes[i]))
	ws_row_count+=1
	ws_col_end=get_column_letter(ws_col_start+len(contenttypes)-1)

	_table_row.append("=SUM("+get_column_letter(ws_col_start)+str(ws_row_count)+":"+ws_col_end+str(ws_row_count)+")")#add calculated cell
	ws.append(_table_row)#second to n line of workbook

def getSubjects():
	cur = conn.cursor()
	cur.execute(get_subjects_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[0])
	return array
def getDatasetSubject(object_id,_table_row):
	global subjects,ws_col_start,ws_row_count
	subject_array=getDatasetSubjects(object_id,0)
	for i in range(len(subjects)):
		_table_row.append(int((subjects[i] in subject_array) == True))
	ws_row_count+=1
	ws_col_end=get_column_letter(ws_col_start+len(subjects)-1)

	_table_row.append("=SUM("+get_column_letter(ws_col_start)+str(ws_row_count)+":"+ws_col_end+str(ws_row_count)+")")#add calculated cell
	ws.append(_table_row)#second to n line of workbook
def getDatasetSubjects(object_id,return_col):
	cur = conn.cursor()
	cur.execute(get_dataset_subjects_sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[return_col])
	return array
def getDatasetFiles(object_id,return_col):
	cur = conn.cursor()
	cur.execute(get_dataset_files_sql.replace("{id}", str(object_id)))
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[return_col])
	return array


def getAffiliations():
	cur = conn.cursor()
	cur.execute(get_affiliations_sql)
	rows = cur.fetchall()
	array=[]
	for row in rows:
		array.append(row[0])
	return array

def getAffiliationsByMonth(affiliations):
	global start_date,end_date,ws_col_start,ws_row_count
	#
	cur = conn.cursor()
	cur.execute(get_users_by_month_affiliations_sql.replace("{start_date}", start_date).replace("{end_date}", end_date))
	rows = cur.fetchall()

	months=[]
	for row in rows:
		months.append(row[0])
	ordered_dates=sorted(set(months))
	#append months to first row
	for date in ordered_dates:
		table_header.append(date.strftime("%b")+" - "+date.strftime("%Y"))
	table_header.append("Totals")

	ws.append(table_header)#first line of workbook
	#
	ws_col_end=get_column_letter(ws_col_start+ len(ordered_dates)-1)
	#
	for a in affiliations:
		_table_row=[0] * (len(ordered_dates))
		_table_row.insert(0, a)
		for row in rows:
			if str(row[2]) == str(a):
				_table_row[ordered_dates.index(row[0])+1]=row[1]
		#Add totals col
		ws_row_count+=1
		_table_row.append("=SUM("+get_column_letter(ws_col_start)+str(ws_row_count)+":"+ws_col_end+str(ws_row_count)+")")#add calculated cell
		ws.append(_table_row)

def addWorkSheetFooter(start_col,calc_col_count,pad):
	#start_col - the position of the fist calculated col 
	#calc_col_count - the number columns to calculate
	#pad - the number of non-calculated cols
	#
	table_row=["Totals"]
	padding = [None] * pad
	array = np.concatenate((table_row, padding), axis=0).tolist()
	for i in range(0, calc_col_count):
		letter=get_column_letter(start_col+i)
		array.append("=SUM("+letter+"2:"+letter+str(ws_row_count)+")")
	ws.append(array)#inject last row with totals
	
#TODO create objects to store varaibles instead of using global variables
#connect to the database
conn = sql_connect.connect()
####create the workbook
wb = Workbook()
ws = wb.active
#
start_date='2014-10-01'
end_date='2017-10-01'
ws_cols_count=12*3#months of year and number years (depending on date ranges)
#
#SHEET 1 ##########################################
needs_table_header=True#Set a Flag which adds the header once withing the called function
level_count=1#used to specify how many levels to drilldown
#
table_header=[]
table_header.append("Top Level")
table_header.append("Category")
table_header.append("Publication Date")
#
ws.title = "Downloads by Dataverse"#Label the worksheet
ws_row_count=1#increments with each added row - used to calculate the totals
ws_col_start=4#the col to start calculateding from used in the last col sum 
getSubDataverses(1,1)
addWorkSheetFooter(4,ws_cols_count+2,2)
#####SHEET 2#######################################
needs_table_header=True
level_count=1
table_header=[]
table_header.append("Root")
table_header.append("Path")
table_header.append("Title")
table_header.append("Name")
table_header.append("Type")
table_header.append("Status")
table_header.append("Publication Date")
table_header.append("Size (KB)")
wb.create_sheet('Downloads by Dataset')
ws = wb["Downloads by Dataset"]
ws_row_count=1
ws_col_start=len(table_header)+1
#
getDatasets("BY_MONTH")
#
addWorkSheetFooter(ws_col_start-1,ws_cols_count+3,ws_col_start-3)

######SHEET 3
needs_table_header=True
level_count=1
#
table_header=[]
table_header.append("Root")
table_header.append("Path")
table_header.append("Title")
table_header.append("Name")
table_header.append("Type")
table_header.append("Status")
table_header.append("Publication Date")
table_header.append("Size (KB)")

wb.create_sheet('File Types')
ws = wb["File Types"]
ws_row_count=1
ws_col_start=len(table_header)+1
contenttypes=getDatasetTypes()
table_header= np.concatenate((table_header, contenttypes), axis=0).tolist()
table_header.append("Totals")
ws.append(table_header)#first line of workbook
getDatasets("BY_CONTENTTYPE")
addWorkSheetFooter(ws_col_start-1,len(contenttypes)+2,ws_col_start-3)

######SHEET 4
needs_table_header=True
level_count=1
#
table_header=[]
table_header.append("Root")
table_header.append("Path")
table_header.append("Title")
table_header.append("Name")
table_header.append("Type")
table_header.append("Status")
table_header.append("Publication Date")
table_header.append("Size (KB)")

wb.create_sheet('Subjects')
ws = wb["Subjects"]
ws_row_count=1
ws_col_start=len(table_header)+1
subjects=getSubjects()
table_header= np.concatenate((table_header, subjects), axis=0).tolist()
table_header.append("Totals")
ws.append(table_header)#first line of workbook
getDatasets("BY_SUBJECT")
addWorkSheetFooter(ws_col_start-1,len(subjects)+2,ws_col_start-3)

###SHEET 5
needs_table_header=True
ws_row_count=1
ws_col_start=2
table_header=[]
table_header.append("Affiliation")
wb.create_sheet("Users by Affiliations")
ws = wb["Users by Affiliations"]
getAffiliationsByMonth(getAffiliations())
addWorkSheetFooter(2,ws_cols_count+2,0)

####
wb.save("Dataverse Usage Report.xlsx")
