from openpyxl.utils import get_column_letter

from metrics.database import database_connection
from metrics.excel import metricsworksheet


class DatasetDownloads(metricsworksheet.MetricsWorkSheet):

    def __init__(self, start_date, end_date, mwb, month_sub_dir):
        self.cursor = None
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = mwb
        self.month_sub_dir = month_sub_dir
        self.worksheet = self.workbook.create_sheet('Downloads by Dataset')
        self.table_header = ['Root', 'Path', 'Title', 'Name', 'Type', 'Status', 'Publication Date', 'Size (KB)']
        self.level_count = 1
        self.ws_col_start = len(self.table_header) + 1
        self.ws_cols_count = 12 * 1 - 1  # months of year and number years (depending on date ranges)
        self.ws_row_count = 1
        self.needs_table_header = True

        with open(self.sql_dir+'get_datasets.sql') as f:
            self.get_datasets_sql = f.read()
        with open(self.sql_dir+'get_dataset_name.sql') as f:
            self.get_dataset_name_sql = f.read()
        with open(self.sql_dir+'dataset_downloads_by_month.sql') as f:
            self.dataset_downloads_by_month_sql = f.read()
        with open(self.sql_dir+'downloads_all_by_dataset_id.sql')as f:
            self.downloads_all_by_dataset_id = f.read()

    def create(self):
        with database_connection.DatabaseConnection() as connection:
            self.cursor = connection.cursor()
            self.get_datasets()
        self.add_worksheet_footer(self.worksheet, self.ws_col_start - 1, self.ws_cols_count + 4, self.ws_col_start - 3,
                                  self.ws_row_count)
        self.save(self.worksheet.title)

    def get_datasets(self):
        self.cursor.execute(self.get_datasets_sql)
        rows = self.cursor.fetchall()
        for row in rows:
            table_row = []  # reset the table row so it can be added with new data
            for r in row:
                table_row.append(r)
            info = self.get_object_name(self.cursor, table_row[0], self.get_dataset_name_sql)
            title = info[0]
            name = info[1]
            status = self.get_dataset_status(self.cursor, table_row[0])
            hierarchy = list(reversed(self.get_hierarchy(self.cursor, table_row[0])))
            del hierarchy[0]  # remove first item from list

            if len(hierarchy) is 0:
                root = "Root"
            else:
                root = hierarchy[0]
                del hierarchy[0]
            array = [root, ">".join(hierarchy), title, name, "Dataset"]
            array.append(status)
            array.append(table_row[2])

            files_array = self.get_dataset_files(self.cursor, table_row[0], 1)
            try:
                array.append(round(int(sum(files_array)) / 1024, 1))  # convert to KB
            except:
                array.append(0)  # when NoneType experienced
            self.get_downloads_by_month_result(table_row[0], array, self.dataset_downloads_by_month_sql,
                                               self.downloads_all_by_dataset_id)

    def get_downloads_by_month_result(self, object_id, _table_row, downloads_by_month_sql, downloads_all_by_id_sql):
        rows = self.get_downloads_by_month(self.cursor, object_id, downloads_by_month_sql, self.start_date,
                                           self.end_date)
        for row in rows:
            # take the first value and add it to the table_header
            # take the second value and add it to the table_row
            if self.needs_table_header:
                self.table_header.append(row[0].strftime("%b") + " - " + row[0].strftime("%Y"))
            r = row[1]
            if r is None:
                r = 0
            _table_row.append(r)
        #
        if self.needs_table_header:
            self.needs_table_header = False
            self.table_header.append("Total")  # add total col
            self.table_header.append("All Downloads")  # add total col
            self.worksheet.append(self.table_header)  # first line of workbook
        self.ws_row_count += 1
        ws_col_end = get_column_letter(self.ws_col_start + self.ws_cols_count)

        _table_row.append(
            "=SUM(" + get_column_letter(self.ws_col_start) + str(self.ws_row_count) + ":" + ws_col_end + str(
                self.ws_row_count) + ")")  # add calculated cell
        _table_row.append(
            self.get_all_downloads(self.cursor, object_id, downloads_all_by_id_sql))  # add all downloads cell
        self.worksheet.append(_table_row)  # second to n line of workbook
