from openpyxl.utils import get_column_letter

from metrics.database.database_connection import DatabaseConnection
from metrics.excel.metricsworksheet import MetricsWorkSheet


class DataverseDownloads(MetricsWorkSheet):

    def __init__(self, start_date, end_date, mwb, month_sub_dir):
        self.cursor = None
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = mwb
        self.month_sub_dir = month_sub_dir
        self.worksheet = self.workbook.create_sheet('Downloads by Dataverse')
        self.table_header = ['Top Level', 'Category', 'Publication Date']
        self.level_count = 1
        self.ws_col_start = len(self.table_header) + 1
        self.ws_cols_count = 12 * 1 - 1  # months of year and number years (depending on date ranges)
        self.ws_row_count = 1
        self.needs_table_header = True

        with open(self.sql_dir+'get_subdataverses.sql') as f:
            self.get_subdataverses = f.read()
        with open(self.sql_dir+'get_dataverse_name.sql') as f:
            self.get_dataverse_name_sql = f.read()
        with open(self.sql_dir+'downloads_by_month.sql') as f:
            self.downloads_by_month_sql = f.read()
        with open(self.sql_dir+'dataset_downloads_by_month.sql') as f:
            self.dataset_downloads_by_month_sql = f.read()
        with open(self.sql_dir+'downloads_all_by_id.sql') as f:
            self.downloads_all_by_id_sql = f.read()

    def create(self):
        with DatabaseConnection() as connection:
            self.cursor = connection.cursor()
            self.get_sub_dataverses(1, 1)
        self.add_worksheet_footer(self.worksheet, 4, self.ws_cols_count + 3, 2, self.ws_row_count)
        self.save(self.worksheet.title)

    def get_sub_dataverses(self, dataverse_id, depth):
        # dataverse_id - the id of the root dataverse
        # depth - the current depth of the root dataverse

        self.cursor.execute(self.get_subdataverses.replace("{dataverse_id}", str(dataverse_id)))
        rows = self.cursor.fetchall()
        for table_row in rows:
            info = self.get_object_name(self.cursor, table_row[0], self.get_dataverse_name_sql)
            array = self.normalize_array(self.level_count, depth, info[1], info[0], table_row[2])
            self.get_downloads_by_month_result(table_row[0], array, self.downloads_by_month_sql,
                                               self.downloads_all_by_id_sql)
            # check if we should go deeper
            if depth < self.level_count:
                self.get_sub_dataverses(table_row[0], depth + 1)

    def get_downloads_by_month_result(self, object_id, _table_row, downloads_by_month_sql, downloads_all_by_id_sql):
        rows = self.get_downloads_by_month(self.cursor, object_id, downloads_by_month_sql, self.start_date,
                                           self.end_date)
        for row in rows:
            # take the first value and add it to the table_header
            # take the second value and add it to the table_row
            if self.needs_table_header:
                self.table_header.append(row[0].strftime("%b") + " - " + row[0].strftime("%Y"))
            r = row[1]
            if r is None:
                r = 0
            _table_row.append(r)
        #
        if self.needs_table_header:
            self.needs_table_header = False
            self.table_header.append("Total")  # add total col
            self.table_header.append("All Downloads")  # add total col
            self.worksheet.append(self.table_header)  # first line of workbook
        self.ws_row_count += 1
        ws_col_end = get_column_letter(self.ws_col_start + self.ws_cols_count)

        _table_row.append(
            "=SUM(" + get_column_letter(self.ws_col_start) + str(self.ws_row_count) + ":" + ws_col_end + str(
                self.ws_row_count) + ")")  # add calculated cell
        _table_row.append(
            self.get_all_downloads(self.cursor, object_id, downloads_all_by_id_sql))  # add all downloads cell
        self.worksheet.append(_table_row)  # second to n line of workbook
