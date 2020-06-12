import csv
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from metrics.config import config


class MetricsWorkSheet:
    start_date = None
    end_date = None
    workbook = None
    month_sub_dir = None
    sql_dir = config.read('configuration').get('sql_dir')
    usage_report = config.read('excel').get('report')

    def __init__(self):
        self._cursor = None
        self._worksheet = None
        self._table_header = None
        self._level_count = None
        self._ws_col_start = None
        self._ws_cols_count = None
        self._ws_row_count = None

    def save(self, sheet_title):
        file = os.path.join(self.month_sub_dir, MetricsWorkSheet.usage_report)
        self.workbook.save(file)
        workbook = load_workbook(file, data_only=True)
        self.create_csv(os.path.join(self.month_sub_dir, sheet_title.replace(' ', '_') + '.csv'), workbook[sheet_title])

    def get_object_name(self, cursor, object_id, sql):
        cursor.execute(sql.replace("{id}", str(object_id)))
        rows = cursor.fetchall()
        return [rows[0][0], rows[0][1]]

    def normalize_array(self, level_count, depth, category, name, publication_date):
        _table_row = [None] * (depth - 1)
        _table_row.append(name)
        # Also append blank cols to the end
        padding = [None] * (level_count - depth)
        array = np.concatenate((_table_row, padding), axis=0).tolist()
        array.append(category)
        array.append(publication_date)
        return array

    def get_downloads_by_month(self, cursor, object_id, sql, start_date, end_date):
        cursor.execute(
            sql.replace("{object_id}", str(object_id)).replace("{start_date}", start_date).replace("{end_date}",
                                                                                                   end_date))
        return cursor.fetchall()

    def get_all_downloads(self, cursor, object_id, sql):
        sql = sql.replace("{object_id}", str(object_id))
        cursor.execute(sql)
        rows = cursor.fetchall()
        for row in rows:
            count = row[0]
        return count

    def add_worksheet_footer(self, ws, start_col, calc_col_count, pad, ws_row_count):
        # start_col - the position of the fist calculated col
        # calc_col_count - the number columns to calculate
        # pad - the number of non-calculated cols
        #
        table_row = ["Totals"]
        padding = [None] * pad
        array = np.concatenate((table_row, padding), axis=0).tolist()
        for i in range(0, calc_col_count):
            letter = get_column_letter(start_col + i)
            array.append("=SUM(" + letter + "2:" + letter + str(ws_row_count) + ")")
        ws.append(array)  # inject last row with totals

    def get_dataset_status(self, cursor, dataset_id):
        with open(self.sql_dir+'get_dataset_status.sql')as f:
            get_dataset_status_sql = f.read()
        cursor.execute(get_dataset_status_sql.replace("{id}", str(dataset_id)))
        rows = cursor.fetchall()
        for r in rows:
            return r[0]

    def get_hierarchy(self, cursor, object_id):
        with open(self.sql_dir+'get_dataset_parents.sql') as f:
            get_dataset_parents_sql = f.read()
        with open(self.sql_dir+'get_dataverse_name.sql') as f:
            get_dataverse_name_sql = f.read()
        cursor.execute(get_dataset_parents_sql.replace("{object_id}", str(object_id)))
        rows = cursor.fetchall()
        array = []
        for r in rows:
            array.append(self.get_object_name(cursor, r[0], get_dataverse_name_sql)[0])
        return array

    def get_dataset_files(self, cursor, object_id, return_col):
        with open(self.sql_dir+'get_dataset_files.sql')as f:
            get_dataset_files_sql = f.read()
        cursor.execute(get_dataset_files_sql.replace("{id}", str(object_id)))
        rows = cursor.fetchall()
        array = []
        for row in rows:
            array.append(row[return_col])
        return array

    def create_csv(self, file_name, sheet):
        with open(file_name, 'w') as f:
            c = csv.writer(f)
            for r in sheet.rows:
                row = []
                for cell in r:
                    row.append(cell.value)
                c.writerow([s for s in row])
