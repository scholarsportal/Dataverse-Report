import numpy as np
from openpyxl.utils import get_column_letter

from metrics.database.database_connection import DatabaseConnection
from metrics.excel.metricsworksheet import MetricsWorkSheet


class Subjects(MetricsWorkSheet):

    def __init__(self, start_date, end_date, mwb, month_sub_dir):
        self.cursor = None
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = mwb
        self.month_sub_dir = month_sub_dir
        self.worksheet = self.workbook.create_sheet('Subjects')
        self.table_header = ['Root', 'Path', 'Title', 'Name', 'Type', 'Status', 'Publication Date', 'Size (KB)']
        self.level_count = 1
        self.ws_col_start = len(self.table_header) + 1
        self.ws_cols_count = 12 * 1 - 1  # months of year and number years (depending on date ranges)
        self.ws_row_count = 1
        self.needs_table_header = True
        self.subjects = None

        with open(self.sql_dir+'get_subjects.sql') as f:
            self.get_subjects_sql = f.read()
        with open(self.sql_dir+'get_dataset_subjects.sql') as f:
            self.get_dataset_subjects_sql = f.read()
        with open(self.sql_dir+'get_datasets.sql') as f:
            self.get_datasets_sql = f.read()
        with open(self.sql_dir+'get_dataset_name.sql') as f:
            self.get_dataset_name_sql = f.read()

    def create(self):
        with DatabaseConnection() as connection:
            self.cursor = connection.cursor()
            self.subjects = self.get_subjects()
            self.table_header = np.concatenate((self.table_header, self.subjects), axis=0).tolist()
            self.table_header.append("Totals")
            self.worksheet.append(self.table_header)  # first line of workbook
            self.get_datasets()
        self.add_worksheet_footer(self.worksheet, self.ws_col_start - 1, len(self.subjects) + 2, self.ws_col_start - 3,
                                  self.ws_row_count)
        self.save(self.worksheet.title)

    def get_subjects(self):
        self.cursor.execute(self.get_subjects_sql)
        rows = self.cursor.fetchall()
        array = []
        for row in rows:
            array.append(row[0])
        return array

    def get_datasets(self):
        self.cursor.execute(self.get_datasets_sql)
        rows = self.cursor.fetchall()
        for row in rows:
            table_row = []  # reset the table row so it can be added with new data
            for r in row:
                table_row.append(r)
            info = self.get_object_name(self.cursor, table_row[0], self.get_dataset_name_sql)
            title = info[0]
            name = info[1]
            status = self.get_dataset_status(self.cursor, table_row[0])
            hierarchy = list(reversed(self.get_hierarchy(self.cursor, table_row[0])))
            del hierarchy[0]  # remove first item from list

            if len(hierarchy) is 0:
                root = "Root"
            else:
                root = hierarchy[0]
                del hierarchy[0]
            array = [root, ">".join(hierarchy), title, name, "Dataset"]
            array.append(status)
            array.append(table_row[2])

            files_array = self.get_dataset_files(self.cursor, table_row[0], 1)
            try:
                array.append(round(int(sum(files_array)) / 1024, 1))  # convert to KB
            except:
                array.append(0)  # when NoneType experienced
            self.get_dataset_subject(table_row[0], array)

    def get_dataset_subject(self, object_id, _table_row):
        subject_array = self.get_dataset_subjects(object_id, 0)
        for i in range(len(self.subjects)):
            _table_row.append(int((self.subjects[i] in subject_array) == True))
        self.ws_row_count += 1
        ws_col_end = get_column_letter(self.ws_col_start + len(self.subjects) - 1)

        _table_row.append(
            "=SUM(" + get_column_letter(self.ws_col_start) + str(self.ws_row_count) + ":" + ws_col_end + str(
                self.ws_row_count) + ")")  # add calculated cell
        self.worksheet.append(_table_row)  # second to n line of workbook

    def get_dataset_subjects(self, object_id, return_col):
        self.cursor.execute(self.get_dataset_subjects_sql.replace("{id}", str(object_id)))
        rows = self.cursor.fetchall()
        array = []
        for row in rows:
            array.append(row[return_col])
        return array
