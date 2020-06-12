from openpyxl.utils import get_column_letter

from metrics.database.database_connection import DatabaseConnection
from metrics.excel.metricsworksheet import MetricsWorkSheet


class UserAffiliation(MetricsWorkSheet):

    def __init__(self, start_date, end_date, mwb, month_sub_dir):
        self.cursor = None
        self.start_date = start_date
        self.end_date = end_date
        self.workbook = mwb
        self.month_sub_dir = month_sub_dir
        self.worksheet = self.workbook.create_sheet('Users by Affiliations')
        self.table_header = ['Affiliation']
        self.ws_col_start = len(self.table_header) + 1
        self.ws_cols_count = 12 * 1 - 1  # months of year and number years (depending on date ranges)
        self.ws_row_count = 1
        self.needs_table_header = True
        self.affiliations = None
        self.ws_col_end = None

        with open(self.sql_dir+'get_affiliations.sql') as f:
            self.get_affiliations_sql = f.read()
        with open(self.sql_dir+'get_users_by_month_affiliations.sql') as f:
            self.get_users_by_month_affiliations_sql = f.read()
        with open(self.sql_dir+'get_all_users_by_affiliation.sql') as f:
            self.get_all_users_by_affiliation_sql = f.read()

    def create(self):
        with DatabaseConnection() as connection:
            self.cursor = connection.cursor()
            rows = self.get_affiliations_by_month(self.start_date, self.end_date)
            months = []
            for row in rows:
                months.append(row[0])
            ordered_dates = sorted(set(months))

            for date in ordered_dates:
                self.table_header.append(date.strftime("%b") + " - " + date.strftime("%Y"))
            self.table_header.append("Totals")
            self.table_header.append("All Time")
            self.worksheet.append(self.table_header)

            affiliations = self.get_affiliations()
            self.ws_col_end = get_column_letter(self.ws_col_start + len(ordered_dates) - 1)

            for a in affiliations:
                _table_row = [0] * (len(ordered_dates))
                _table_row.insert(0, a)
                for row in rows:
                    if str(row[2]) == str(a):
                        _table_row[ordered_dates.index(row[0]) + 1] = row[1]
                # Add totals col
                self.ws_row_count += 1
                _table_row.append("=SUM(" + get_column_letter(self.ws_col_start) + str(
                    self.ws_row_count) + ":" + self.ws_col_end + str(
                    self.ws_row_count) + ")")  # add calculated cell
                _table_row.append(self.get_all_users_by_affiliation(a))
                self.worksheet.append(_table_row)
            self.add_worksheet_footer(self.worksheet, 2, self.ws_cols_count + 3, 0, self.ws_row_count)
            self.save(self.worksheet.title)

    def get_affiliations(self):
        self.cursor.execute(self.get_affiliations_sql)
        rows = self.cursor.fetchall()
        array = []
        for row in rows:
            array.append(row[0])
        return array

    def get_affiliations_by_month(self, start_date, end_date):
        self.cursor.execute(
            self.get_users_by_month_affiliations_sql.replace("{start_date}", start_date).replace("{end_date}",
                                                                                                 end_date))
        return self.cursor.fetchall()

    def get_all_users_by_affiliation(self, affiliation):
        sql = self.get_all_users_by_affiliation_sql.replace("{affiliation}", str(affiliation).replace("'", "''"))
        self.cursor.execute(sql)
        rows = self.cursor.fetchall()
        for row in rows:
            count = row[0]
        return count
